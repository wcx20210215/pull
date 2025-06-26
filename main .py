import matplotlib.pyplot as plt
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import seaborn as sns
import streamlit as st
from utils import (dataframe_agent, get_session_id, display_conversation_history, 
                   display_popular_questions, get_memory_stats, memory_manager, 
                   streaming_handler, clear_session_memory, clear_all_memory)
from datetime import datetime
import io
import json
import asyncio
import time
import os
import sqlite3
from functools import wraps
from concurrent.futures import ThreadPoolExecutor

# 缓存机制
@st.cache_data(ttl=3600)  # 缓存1小时
def load_cached_data(file_path, file_type, sheet_name=None):
    """缓存数据加载"""
    if file_type == "xlsx":
        return pd.read_excel(file_path, sheet_name=sheet_name)
    elif file_type == "csv":
        return pd.read_csv(file_path)
    elif file_type == "json":
        return pd.read_json(file_path)
    elif file_type == "tsv":
        return pd.read_csv(file_path, sep='\t')
    else:
        raise ValueError(f"不支持的文件类型: {file_type}")

@st.cache_data(ttl=1800)  # 缓存30分钟
def cached_statistical_analysis(df_hash, analysis_type, column=None):
    """缓存统计分析结果"""
    # 这里实际上需要传入DataFrame，但为了缓存我们使用hash
    pass

def async_task_wrapper(func):
    """异步任务包装器"""
    @wraps(func)
    def wrapper(*args, **kwargs):
        with ThreadPoolExecutor(max_workers=2) as executor:
            future = executor.submit(func, *args, **kwargs)
            return future
    return wrapper

@async_task_wrapper
def heavy_computation(df, operation):
    """重计算任务的异步处理"""
    time.sleep(0.1)  # 模拟计算时间
    if operation == "correlation":
        return df.select_dtypes(include=[np.number]).corr()
    elif operation == "describe":
        return df.describe()
    return None

def create_chart(input_data, chart_type):
    """生成统计图表 - 增强版"""
    df_data = pd.DataFrame(
        data={
            "x": input_data["columns"],
            "y": input_data["data"]
        }
    ).set_index("x")
    
    if chart_type == "bar":
        fig = px.bar(
            x=input_data["columns"], 
            y=input_data["data"], 
            title="📊 柱状图分析", 
            color=input_data["data"],
            color_continuous_scale="viridis",
            labels={'x': '类别', 'y': '数值'}
        )
        fig.update_layout(
            title_font_size=16,
            xaxis_title_font_size=14,
            yaxis_title_font_size=14,
            showlegend=False
        )
        st.plotly_chart(fig, use_container_width=True)
        
    elif chart_type == "line":
        fig = px.line(
            x=input_data["columns"], 
            y=input_data["data"], 
            title="📈 趋势分析", 
            markers=True,
            line_shape='spline'
        )
        fig.update_traces(
            line=dict(width=3),
            marker=dict(size=8)
        )
        fig.update_layout(
            title_font_size=16,
            xaxis_title="时间/类别",
            yaxis_title="数值",
            xaxis_title_font_size=14,
            yaxis_title_font_size=14
        )
        st.plotly_chart(fig, use_container_width=True)
        
    elif chart_type == "pie":
        fig = px.pie(
            values=input_data["data"], 
            names=input_data["columns"], 
            title="🥧 饼图分析",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        fig.update_traces(textposition='inside', textinfo='percent+label')
        fig.update_layout(title_font_size=16)
        st.plotly_chart(fig, use_container_width=True)
        
    elif chart_type == "table":
        # 表格数据展示
        table_df = pd.DataFrame({
            '项目': input_data["columns"],
            '数值': input_data["data"]
        })
        st.markdown("#### 📋 数据表格")
        st.dataframe(
            table_df, 
            use_container_width=True,
            hide_index=True
        )
        
        # 添加统计摘要
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("总计", f"{sum(input_data['data']):.2f}")
        with col2:
            st.metric("平均值", f"{np.mean(input_data['data']):.2f}")
        with col3:
            st.metric("最大值", f"{max(input_data['data']):.2f}")
        with col4:
            st.metric("最小值", f"{min(input_data['data']):.2f}")

def display_data_overview(df):
    """显示数据概览"""
    st.subheader("📊 数据概览")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("总行数", len(df))
    with col2:
        st.metric("总列数", len(df.columns))
    with col3:
        st.metric("数值列", len(df.select_dtypes(include=[np.number]).columns))
    with col4:
        st.metric("缺失值", df.isnull().sum().sum())
    
    # 数据类型分布
    st.subheader("📋 列信息")
    col_info = pd.DataFrame({
        '列名': df.columns,
        '数据类型': df.dtypes.astype(str),
        '非空值': df.count(),
        '缺失值': df.isnull().sum(),
        '唯一值': df.nunique()
    })
    st.dataframe(col_info, use_container_width=True)

def data_cleaning_section(df):
    """数据清洗功能 - 增强版"""
    st.subheader("🧹 数据清洗与预处理")
    
    # 添加数据质量评估
    with st.expander("📊 数据质量评估", expanded=False):
        col1, col2, col3 = st.columns(3)
        with col1:
            missing_ratio = (df.isnull().sum().sum() / (len(df) * len(df.columns))) * 100
            st.metric("缺失值比例", f"{missing_ratio:.2f}%")
        with col2:
            duplicate_ratio = (df.duplicated().sum() / len(df)) * 100
            st.metric("重复值比例", f"{duplicate_ratio:.2f}%")
        with col3:
            numeric_ratio = (len(df.select_dtypes(include=[np.number]).columns) / len(df.columns)) * 100
            st.metric("数值列比例", f"{numeric_ratio:.2f}%")
    
    cleaning_option = st.selectbox(
        "选择清洗操作",
        ["查看缺失值", "删除缺失值", "填充缺失值", "删除重复值", "数据类型转换", 
         "异常值检测", "数据标准化", "数据去重增强", "列重命名"]
    )
    
    if cleaning_option == "查看缺失值":
        missing_data = df.isnull().sum()
        missing_data = missing_data[missing_data > 0]
        if len(missing_data) > 0:
            st.write("缺失值统计:")
            st.bar_chart(missing_data)
        else:
            st.success("数据中没有缺失值！")
    
    elif cleaning_option == "删除缺失值":
        if st.button("删除包含缺失值的行"):
            cleaned_df = df.dropna()
            st.success(f"删除了 {len(df) - len(cleaned_df)} 行包含缺失值的数据")
            st.session_state["df"] = cleaned_df
            st.rerun()
    
    elif cleaning_option == "填充缺失值":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            col_to_fill = st.selectbox("选择要填充的列", numeric_cols)
            fill_method = st.selectbox("填充方法", ["均值", "中位数", "众数"])
            
            if st.button("执行填充"):
                if fill_method == "均值":
                    df[col_to_fill].fillna(df[col_to_fill].mean(), inplace=True)
                elif fill_method == "中位数":
                    df[col_to_fill].fillna(df[col_to_fill].median(), inplace=True)
                elif fill_method == "众数":
                    df[col_to_fill].fillna(df[col_to_fill].mode()[0], inplace=True)
                
                st.session_state["df"] = df
                st.success(f"已用{fill_method}填充 {col_to_fill} 列的缺失值")
                st.rerun()
    
    elif cleaning_option == "删除重复值":
        duplicates = df.duplicated().sum()
        st.write(f"发现 {duplicates} 行重复数据")
        if duplicates > 0 and st.button("删除重复值"):
            df_cleaned = df.drop_duplicates()
            st.session_state["df"] = df_cleaned
            st.success(f"删除了 {duplicates} 行重复数据")
            st.rerun()
    
    elif cleaning_option == "异常值检测":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            selected_col = st.selectbox("选择要检测异常值的列", numeric_cols)
            method = st.selectbox("检测方法", ["IQR方法", "Z-Score方法"])
            
            if method == "IQR方法":
                Q1 = df[selected_col].quantile(0.25)
                Q3 = df[selected_col].quantile(0.75)
                IQR = Q3 - Q1
                lower_bound = Q1 - 1.5 * IQR
                upper_bound = Q3 + 1.5 * IQR
                outliers = df[(df[selected_col] < lower_bound) | (df[selected_col] > upper_bound)]
                
            else:  # Z-Score方法
                z_scores = np.abs((df[selected_col] - df[selected_col].mean()) / df[selected_col].std())
                outliers = df[z_scores > 3]
            
            st.write(f"检测到 {len(outliers)} 个异常值")
            if len(outliers) > 0:
                st.dataframe(outliers, use_container_width=True)
                if st.button("删除异常值"):
                    df_cleaned = df.drop(outliers.index)
                    st.session_state["df"] = df_cleaned
                    st.success(f"删除了 {len(outliers)} 个异常值")
                    st.rerun()
        else:
            st.warning("没有数值列可以进行异常值检测")
    
    elif cleaning_option == "数据标准化":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            selected_cols = st.multiselect("选择要标准化的列", numeric_cols)
            method = st.selectbox("标准化方法", ["Z-Score标准化", "Min-Max标准化"])
            
            if selected_cols and st.button("执行标准化"):
                df_copy = df.copy()
                for col in selected_cols:
                    if method == "Z-Score标准化":
                        df_copy[col] = (df[col] - df[col].mean()) / df[col].std()
                    else:  # Min-Max标准化
                        df_copy[col] = (df[col] - df[col].min()) / (df[col].max() - df[col].min())
                
                st.session_state["df"] = df_copy
                st.success(f"已对 {len(selected_cols)} 列执行{method}")
                st.rerun()
        else:
            st.warning("没有数值列可以进行标准化")
    
    elif cleaning_option == "列重命名":
        st.write("当前列名:")
        new_names = {}
        for col in df.columns:
            new_name = st.text_input(f"重命名 '{col}'", value=col, key=f"rename_{col}")
            if new_name != col:
                new_names[col] = new_name
        
        if new_names and st.button("应用重命名"):
            df_renamed = df.rename(columns=new_names)
            st.session_state["df"] = df_renamed
            st.success(f"已重命名 {len(new_names)} 列")
            st.rerun()

def statistical_analysis(df):
    """统计分析功能 - 增强版"""
    st.subheader("📈 统计分析")
    
    analysis_type = st.selectbox(
        "选择分析类型",
        ["描述性统计", "相关性分析", "分布分析", "回归分析", "聚类分析", "时间序列分析"]
    )
    
    if analysis_type == "描述性统计":
        numeric_df = df.select_dtypes(include=[np.number])
        if len(numeric_df.columns) > 0:
            st.write("数值列统计摘要:")
            st.dataframe(numeric_df.describe(), use_container_width=True)
        else:
            st.warning("没有数值列可以进行统计分析")
    
    elif analysis_type == "相关性分析":
        numeric_df = df.select_dtypes(include=[np.number])
        if len(numeric_df.columns) > 1:
            corr_matrix = numeric_df.corr()
            fig = px.imshow(corr_matrix, 
                          text_auto=True, 
                          aspect="auto",
                          title="相关性热力图")
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("需要至少2个数值列才能进行相关性分析")
    
    elif analysis_type == "分布分析":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            selected_col = st.selectbox("选择要分析的列", numeric_cols)
            
            col1, col2 = st.columns(2)
            with col1:
                # 直方图
                fig_hist = px.histogram(df, x=selected_col, title=f"{selected_col} 分布直方图")
                st.plotly_chart(fig_hist, use_container_width=True)
            
            with col2:
                # 箱线图
                fig_box = px.box(df, y=selected_col, title=f"{selected_col} 箱线图")
                st.plotly_chart(fig_box, use_container_width=True)
    
    elif analysis_type == "回归分析":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) >= 2:
            col1, col2 = st.columns(2)
            with col1:
                x_col = st.selectbox("选择自变量(X)", numeric_cols)
            with col2:
                y_col = st.selectbox("选择因变量(Y)", numeric_cols)
            
            if x_col != y_col:
                # 计算相关系数
                correlation = df[x_col].corr(df[y_col])
                st.metric("相关系数", f"{correlation:.4f}")
                
                # 绘制散点图和回归线
                fig = px.scatter(df, x=x_col, y=y_col, 
                               title=f"{x_col} vs {y_col} 回归分析",
                               trendline="ols")
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("需要至少2个数值列才能进行回归分析")
    
    elif analysis_type == "聚类分析":
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) >= 2:
            selected_cols = st.multiselect("选择用于聚类的列", numeric_cols, default=list(numeric_cols[:2]))
            n_clusters = st.slider("聚类数量", 2, 10, 3)
            
            if len(selected_cols) >= 2 and st.button("执行聚类分析"):
                from sklearn.cluster import KMeans
                from sklearn.preprocessing import StandardScaler
                
                # 数据预处理
                data_for_clustering = df[selected_cols].dropna()
                scaler = StandardScaler()
                scaled_data = scaler.fit_transform(data_for_clustering)
                
                # K-means聚类
                kmeans = KMeans(n_clusters=n_clusters, random_state=42)
                clusters = kmeans.fit_predict(scaled_data)
                
                # 可视化结果
                if len(selected_cols) == 2:
                    fig = px.scatter(x=data_for_clustering.iloc[:, 0], 
                                   y=data_for_clustering.iloc[:, 1],
                                   color=clusters,
                                   title=f"K-means聚类结果 (k={n_clusters})",
                                   labels={'x': selected_cols[0], 'y': selected_cols[1]})
                    st.plotly_chart(fig, use_container_width=True)
                
                # 显示聚类统计
                cluster_stats = pd.DataFrame({
                    '聚类': range(n_clusters),
                    '样本数': [sum(clusters == i) for i in range(n_clusters)]
                })
                st.dataframe(cluster_stats, use_container_width=True)
        else:
            st.warning("需要至少2个数值列才能进行聚类分析")
    
    elif analysis_type == "时间序列分析":
        date_cols = df.select_dtypes(include=['datetime64', 'object']).columns
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        
        if len(date_cols) > 0 and len(numeric_cols) > 0:
            date_col = st.selectbox("选择时间列", date_cols)
            value_col = st.selectbox("选择数值列", numeric_cols)
            
            try:
                # 尝试转换为日期类型
                df_ts = df.copy()
                df_ts[date_col] = pd.to_datetime(df_ts[date_col])
                df_ts = df_ts.sort_values(date_col)
                
                # 绘制时间序列图
                fig = px.line(df_ts, x=date_col, y=value_col, 
                            title=f"{value_col} 时间序列分析")
                st.plotly_chart(fig, use_container_width=True)
                
                # 基本统计
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("趋势", "上升" if df_ts[value_col].iloc[-1] > df_ts[value_col].iloc[0] else "下降")
                with col2:
                    volatility = df_ts[value_col].std()
                    st.metric("波动性", f"{volatility:.2f}")
                with col3:
                    growth_rate = ((df_ts[value_col].iloc[-1] / df_ts[value_col].iloc[0]) - 1) * 100
                    st.metric("总增长率", f"{growth_rate:.2f}%")
                    
            except Exception as e:
                st.error(f"时间序列分析失败: {str(e)}")
        else:
            st.warning("需要至少1个时间列和1个数值列才能进行时间序列分析")

def advanced_visualization(df):
    """高级可视化功能"""
    st.subheader("🎨 高级可视化")
    
    chart_type = st.selectbox(
        "选择图表类型",
        ["散点图", "饼图", "热力图", "小提琴图"]
    )
    
    numeric_cols = df.select_dtypes(include=[np.number]).columns
    categorical_cols = df.select_dtypes(include=['object']).columns
    
    if chart_type == "散点图" and len(numeric_cols) >= 2:
        col1, col2 = st.columns(2)
        with col1:
            x_col = st.selectbox("X轴", numeric_cols)
        with col2:
            y_col = st.selectbox("Y轴", numeric_cols)
        
        color_col = None
        if len(categorical_cols) > 0:
            color_col = st.selectbox("颜色分组（可选）", ["无"] + list(categorical_cols))
            color_col = None if color_col == "无" else color_col
        
        fig = px.scatter(df, x=x_col, y=y_col, color=color_col, 
                        title=f"{x_col} vs {y_col} 散点图")
        st.plotly_chart(fig, use_container_width=True)
    
    elif chart_type == "饼图" and len(categorical_cols) > 0:
        cat_col = st.selectbox("选择分类列", categorical_cols)
        value_counts = df[cat_col].value_counts()
        
        fig = px.pie(values=value_counts.values, names=value_counts.index, 
                    title=f"{cat_col} 分布饼图")
        st.plotly_chart(fig, use_container_width=True)
    
    elif chart_type == "热力图" and len(numeric_cols) > 1:
        selected_cols = st.multiselect("选择要显示的列", numeric_cols, default=list(numeric_cols[:5]))
        if selected_cols:
            corr_data = df[selected_cols].corr()
            fig = px.imshow(corr_data, text_auto=True, aspect="auto", 
                          title="选定列相关性热力图")
            st.plotly_chart(fig, use_container_width=True)
    
    else:
        st.warning("当前数据不支持所选图表类型，请检查数据类型和列数量")

def export_data_section(df):
    """数据导出功能"""
    st.subheader("💾 数据导出")
    
    export_format = st.selectbox("选择导出格式", ["CSV", "Excel", "JSON"])
    
    if export_format == "CSV":
        csv = df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="下载CSV文件",
            data=csv,
            file_name=f"processed_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )
    
    elif export_format == "Excel":
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='数据', index=False)
        
        st.download_button(
            label="下载Excel文件",
            data=output.getvalue(),
            file_name=f"processed_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    elif export_format == "JSON":
        json_data = df.to_json(orient='records', force_ascii=False, indent=2)
        st.download_button(
            label="下载JSON文件",
            data=json_data,
            file_name=f"processed_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json"
        )
# 页面配置
st.set_page_config(
    page_title="小脑瓜数据分析智能体",
    page_icon="🧠",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 主标题
st.title("🧠 小脑瓜数据分析智能体")
st.markdown("---")

# 侧边栏功能选择
with st.sidebar:
    st.header("🎯 功能选择")
    function_choice = st.selectbox(
        "选择功能模块",
        ["数据上传", "数据概览", "数据清洗", "统计分析", "高级可视化", "AI问答", "数据导出"]
    )
    
    st.markdown("---")
    st.markdown("### 📋 使用说明")
    st.markdown("""
    1. 首先上传数据文件
    2. 选择相应功能模块
    3. 根据需要进行数据分析
    4. 导出处理结果
    """)

# 数据上传区域 - 增强版
if function_choice == "数据上传" or "df" not in st.session_state:
    st.subheader("📁 数据上传")
    
    # 文件格式选择
    option = st.radio(
        "请选择数据文件类型:", 
        ("Excel", "CSV", "JSON", "TSV", "Parquet")
    )
    
    # 文件类型映射
    file_type_map = {
        "Excel": "xlsx",
        "CSV": "csv", 
        "JSON": "json",
        "TSV": "tsv",
        "Parquet": "parquet"
    }
    
    file_type = file_type_map[option]
    
    # 文件上传提示
    upload_help = {
        "Excel": "支持 .xlsx 和 .xls 格式，可选择工作表",
        "CSV": "支持逗号分隔的文本文件",
        "JSON": "支持标准JSON格式的数据文件", 
        "TSV": "支持制表符分隔的文本文件",
        "Parquet": "支持高效的列式存储格式"
    }
    
    st.info(f"💡 {upload_help[option]}")
    
    # 文件大小限制提示
    st.caption("📏 最大文件大小: 200MB")
    
    data = st.file_uploader(
        f"上传你的{option}数据文件", 
        type=[file_type] if file_type != "xlsx" else ["xlsx", "xls"]
    )
    
    if data:
        try:
            with st.spinner("🔄 正在加载数据..."):
                if file_type == "xlsx" or option == "Excel":
                    wb = openpyxl.load_workbook(data)
                    if len(wb.sheetnames) > 1:
                        sheet_option = st.radio(
                            label="请选择要加载的工作表：", 
                            options=wb.sheetnames
                        )
                    else:
                        sheet_option = wb.sheetnames[0]
                    st.session_state["df"] = pd.read_excel(data, sheet_name=sheet_option)
                    
                elif file_type == "csv":
                    # CSV编码检测
                    encoding = st.selectbox(
                        "选择文件编码", 
                        ["utf-8", "gbk", "gb2312", "utf-8-sig"],
                        index=0
                    )
                    st.session_state["df"] = pd.read_csv(data, encoding=encoding)
                    
                elif file_type == "json":
                    st.session_state["df"] = pd.read_json(data)
                    
                elif file_type == "tsv":
                    encoding = st.selectbox(
                        "选择文件编码", 
                        ["utf-8", "gbk", "gb2312", "utf-8-sig"],
                        index=0
                    )
                    st.session_state["df"] = pd.read_csv(data, sep='\t', encoding=encoding)
                    
                elif file_type == "parquet":
                    st.session_state["df"] = pd.read_parquet(data)
            
            st.success("✅ 数据上传成功！")
            
            # 数据基本信息
            df = st.session_state["df"]
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("行数", len(df))
            with col2:
                st.metric("列数", len(df.columns))
            with col3:
                st.metric("内存使用", f"{df.memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB")
            with col4:
                st.metric("数据类型", len(df.dtypes.unique()))
            
            with st.expander("🔍 预览原始数据", expanded=True):
                st.dataframe(st.session_state["df"], use_container_width=True)
                
        except Exception as e:
            st.error(f"❌ 数据加载失败: {str(e)}")
            st.info("💡 请检查文件格式是否正确，或尝试其他编码方式")

# 功能模块展示
if "df" in st.session_state:
    df = st.session_state["df"]
    
    if function_choice == "数据概览":
        display_data_overview(df)
    
    elif function_choice == "数据清洗":
        data_cleaning_section(df)
    
    elif function_choice == "统计分析":
        statistical_analysis(df)
    
    elif function_choice == "高级可视化":
        advanced_visualization(df)
    
    elif function_choice == "数据导出":
        export_data_section(df)
    
    elif function_choice == "AI问答":
        st.subheader("🤖 AI智能问答")
        
        # 获取会话ID
        session_id = get_session_id()
        
        # 记忆统计信息
        memory_stats = get_memory_stats(df)
        
        # 顶部统计卡片
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("💬 本次对话", memory_stats["session_count"])
        with col2:
            st.metric("⚡ 快速回答", memory_stats["quick_answers_count"])
        with col3:
            st.metric("📊 总对话数", memory_stats["total_conversations"])
        with col4:
            st.metric("⏱️ 平均响应", f"{memory_stats['avg_response_time']:.2f}s")
        
        # 数据概览卡片
        with st.expander("📊 当前数据概览", expanded=False):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("数据行数", len(df))
            with col2:
                st.metric("数据列数", len(df.columns))
            with col3:
                numeric_cols = len(df.select_dtypes(include=[np.number]).columns)
                st.metric("数值列", numeric_cols)
            with col4:
                categorical_cols = len(df.select_dtypes(include=['object']).columns)
                st.metric("文本列", categorical_cols)
            
            st.write("**列名预览:**", ", ".join(df.columns[:10].tolist()) + ("..." if len(df.columns) > 10 else ""))
        
        # 侧边栏：对话历史和热门问题
        with st.sidebar:
            st.markdown("---")
            
            # 对话历史
            with st.expander("📚 对话历史", expanded=False):
                display_conversation_history(df, limit=3)
            
            # 热门问题
            with st.expander("🔥 热门问题", expanded=False):
                display_popular_questions(df, limit=3)
            
            # 清除记忆按钮
            col1, col2 = st.columns(2)
            with col1:
                if st.button("🗑️ 清除会话记忆"):
                    deleted_count = clear_session_memory(session_id)
                    st.success(f"已清除 {deleted_count} 条会话记忆！")
                    st.rerun()
            with col2:
                if st.button("💥 清除所有记忆"):
                    if st.session_state.get('confirm_clear_all', False):
                        clear_all_memory()
                        st.success("所有记忆已清除！")
                        st.session_state.confirm_clear_all = False
                        st.rerun()
                    else:
                        st.session_state.confirm_clear_all = True
                        st.warning("再次点击确认清除所有记忆")
                        st.rerun()
        
        # 快速问题模板 - 增强版
        st.markdown("#### 💡 智能问题模板")
        
        # 分类问题模板
        template_categories = {
            "📊 基础统计": [
                "显示数据的基本统计信息",
                "计算数值列的平均值和标准差",
                "找出缺失值最多的列",
                "显示数据类型分布"
            ],
            "🔍 数据探索": [
                "找出数值最大的前5行数据",
                "显示各类别的分布情况", 
                "找出异常值或离群点",
                "计算数值列之间的相关性"
            ],
            "📈 可视化分析": [
                "生成销售额的柱状图",
                "创建时间序列趋势图",
                "绘制相关性热力图",
                "制作分类数据的饼图"
            ],
            "🎯 高级分析": [
                "进行聚类分析并可视化结果",
                "执行回归分析找出关联关系",
                "预测未来趋势",
                "识别数据中的模式和规律"
            ]
        }
        
        # 选择问题类别
        selected_category = st.selectbox(
            "选择问题类别", 
            ["自定义问题"] + list(template_categories.keys())
        )
        
        # 检查是否选择了热门问题
        initial_query = ""
        if "selected_question" in st.session_state:
            initial_query = st.session_state.selected_question
            del st.session_state.selected_question
        
        if selected_category != "自定义问题":
            selected_template = st.selectbox(
                "选择具体问题", 
                template_categories[selected_category]
            )
            query = st.text_area(
                "💬 请输入你关于数据集的问题或可视化需求：",
                value=initial_query or selected_template,
                height=100,
                help="你可以修改模板问题或直接使用"
            )
        else:
            query = st.text_area(
                "💬 请输入你关于数据集的问题或可视化需求：",
                value=initial_query,
                placeholder="例如：显示销售额最高的前5个地区的柱状图，并分析其趋势",
                height=100,
                help="支持中文问题，可以要求生成图表、表格或进行数据分析"
            )
        
        # 高级选项
        with st.expander("⚙️ 高级选项", expanded=False):
            col1, col2, col3 = st.columns(3)
            with col1:
                response_format = st.selectbox(
                    "期望的回答格式",
                    ["智能选择", "纯文字", "表格数据", "图表可视化", "综合分析"]
                )
            with col2:
                analysis_depth = st.selectbox(
                    "分析深度",
                    ["标准", "详细", "简洁"]
                )
            with col3:
                enable_streaming = st.checkbox(
                    "🌊 流式输出",
                    value=True,
                    help="启用流式输出可以实时看到AI的思考过程"
                )
            
            # 记忆选项
            col1, col2 = st.columns(2)
            with col1:
                use_memory = st.checkbox(
                    "🧠 启用记忆",
                    value=True,
                    help="启用记忆功能可以快速回答相同问题"
                )
            with col2:
                show_cache_info = st.checkbox(
                    "📊 显示缓存信息",
                    value=False,
                    help="显示是否使用了缓存回答"
                )
        
        # 生成回答按钮
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            button = st.button("🚀 生成AI分析", type="primary", use_container_width=True)
        
        if query and button:
            # 构建增强的查询
            enhanced_query = query
            if response_format != "智能选择":
                enhanced_query += f" (请以{response_format}的形式回答)"
            if analysis_depth != "标准":
                enhanced_query += f" (分析深度：{analysis_depth})"
            
            # 检查是否有快速回答
            data_hash = memory_manager.get_data_hash(df)
            quick_answer = None
            if use_memory:
                quick_answer = memory_manager.get_quick_answer(enhanced_query, data_hash)
            
            # 显示缓存信息
            if show_cache_info and quick_answer and quick_answer["found"]:
                st.info(f"⚡ 找到快速回答！已使用 {quick_answer['hit_count']} 次")
            
            # 创建结果容器
            result_container = st.container()
            
            with st.spinner("🤔 AI正在深度分析中，请稍等..."):
                start_time = time.time()
                
                # 准备流式输出容器
                stream_container = None
                if enable_streaming and not (quick_answer and quick_answer["found"]):
                    stream_container = st.empty()
                
                # 调用增强的dataframe_agent
                result = dataframe_agent(
                    df=df, 
                    query=enhanced_query,
                    session_id=session_id,
                    use_cache=use_memory,
                    enable_streaming=enable_streaming,
                    stream_container=stream_container
                )
                
                end_time = time.time()
                
                # 清除流式输出容器
                if stream_container:
                    stream_container.empty()
                
                with result_container:
                    st.markdown("### 🎯 AI分析结果")
                    
                    # 显示处理时间和缓存状态
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.caption(f"⏱️ 分析耗时: {end_time - start_time:.2f}秒")
                    with col2:
                        if quick_answer and quick_answer["found"]:
                            st.caption("⚡ 快速回答")
                        else:
                            st.caption("🆕 新分析")
                    with col3:
                        if use_memory:
                            st.caption("🧠 已保存到记忆")
                    
                    if "answer" in result:
                        st.success(result["answer"])
                    
                    if "table" in result:
                        st.markdown("#### 📊 数据表格")
                        result_df = pd.DataFrame(result["table"]["data"],
                                               columns=result["table"]["columns"])
                        st.dataframe(result_df, use_container_width=True)
                        
                        # 添加导出选项
                        csv = result_df.to_csv(index=False, encoding='utf-8-sig')
                        st.download_button(
                            label="📥 下载表格数据",
                            data=csv,
                            file_name=f"analysis_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv"
                        )
                    
                    if "bar" in result:
                        st.markdown("#### 📊 柱状图分析")
                        create_chart(result["bar"], "bar")
                    
                    if "line" in result:
                        st.markdown("#### 📈 趋势分析")
                        create_chart(result["line"], "line")
                    
                    if "pie" in result:
                        st.markdown("#### 🥧 饼图分析")
                        create_chart(result["pie"], "pie")
                
                # 添加反馈机制
                st.markdown("---")
                st.markdown("#### 💭 分析反馈")
                col1, col2, col3 = st.columns(3)
                with col1:
                    if st.button("👍 满意"):
                        st.success("感谢您的反馈！")
                with col2:
                    if st.button("👎 不满意"):
                        st.info("我们会继续改进，请尝试更具体的问题描述")
                with col3:
                    if st.button("🔄 重新分析"):
                        st.rerun()

else:
    if function_choice != "数据上传":
        st.warning("⚠️ 请先上传数据文件才能使用此功能")
        st.info("👈 请在侧边栏选择'数据上传'功能")